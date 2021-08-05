from openpyxl import Workbook

myFile = r'\\vwoachsfile01\assembly\Departments\Assembly Launch\Battery Plant Status\Pilot Hall\IOL_EOL_Leak Docs\PVS\288 TX6PVS0340_2021_06_16_180429.ini'
input_file = open(myFile,'r')

#Variable Setup
#Master dictionary
Master = {}

#Loop through each line in file
while input_file:
    # Read next line
    currentLine = input_file.readline()
    #Check if EOF
    if currentLine == "":
        break
    #Identify header (will be separate dictionary entry in master dictionary
    if currentLine[0] == "[":
        header = currentLine.strip()
        #Create a blank dictionary to temporarily save to before adding to master
        myDict = {}
        dictName = currentLine

        #Index to next line
        currentLine = input_file.readline()

        while currentLine.strip() != "":

            #If there is a ";" in the string
            if ";" in currentLine.strip():
                data = currentLine[currentLine.find('"') + 1:].strip()
                data = data[:-1]
                # Create list of data/value pair
                myData = data.split(";")
                myDict[myData[0]] = myData[1]
            # If there is NOT a ";" in the string
            else:
                #      0 = "IO"
                data = currentLine
                myData = currentLine.split("=")
                myDict[myData[0].strip(" ")] = myData[1].strip().strip(" ")

            #Index to the next line at end of loop
            currentLine = input_file.readline()

        Master[header] = myDict

#Print each dictionary
#for (key,value) in Master.items():
    #print(key, value)

###############################################
#Print to Excel

wb = Workbook()
sheet = wb.active
count = 0

#General Data
battery = Master["[Pruefling]"]["Seriennummer"][7:15]
partNum = Master["[Pruefling]"]["Teilenummer"]
batType = Master["[Pruefling]"]["Batteriesystemtyp"]

#PreCheck
i_DTC_Count = Master["[SW_Steuergeraete_Allgemeine Informationen]"]["uPruefungszaehlerIst"]

#Software Setpoint Check
initBCMe = Master["[SW_Steuergeraete_Pruefergebnisse]"]["SW_BMCe_Get"]
initCMC1 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["SW_CMC1_Get"]
initCMC2 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["SW_CMC2_Get"]
initCMC3 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["SW_CMC3_Get"]
initBL_BCMe = Master["[SW_Steuergeraete_Pruefergebnisse]"]["BL_BMCe_Get"]
initBL_CMC1 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["BL_CMC1_Get"]
initBL_CMC2 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["BL_CMC2_Get"]
initBL_CMC3 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["BL_CMC3_Get"]

#Default parameters set
batConfigPset = Master["[SW_Steuergeraete_Pruefergebnisse]"]["DatensatzVersion_Batteriekonfig_Get"]
batVehiclePset = Master["[SW_Steuergeraete_Pruefergebnisse]"]["DatensatzVersion_Fahrzeug_Get"]
batVehiclePset = Master["[SW_Steuergeraete_Pruefergebnisse]"]["DatensatzVersion_Fahrzeug_Get"]
batTargetMarketPset = Master["[SW_Steuergeraete_Pruefergebnisse]"]["DatensatzVersion_Zielmarkt_Get"]
batConfigNamePset = Master["[SW_Steuergeraete_Pruefergebnisse]"]["DatensatzName_Batteriekonfig_Get"]
batThermoPset = Master["[SW_Steuergeraete_Pruefergebnisse]"]["DatensatzName_Thermo_Get"]
batNameVehiclePset = Master["[SW_Steuergeraete_Pruefergebnisse]"]["DatensatzName_Fahrzeug_Get"]
batNameTargetMarket = Master["[SW_Steuergeraete_Pruefergebnisse]"]["DatensatzName_Zielmarkt_Get"]

#Hardware PN
HW_NumCMC1 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F1A3_CMC1_Get"]
HW_NumCMC2 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F1A3_CMC2_Get"]
HW_NumCMC3 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F1A3_CMC3_Get"]
serial_CMC1 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["SerialNr_CMC1_Get"]
serial_CMC2 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["SerialNr_CMC2_Get"]
serial_CMC3 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["SerialNr_CMC3_Get"]
HW_NumBMCe = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F1A3_BMCe_Get"]
HW_serial_BMCe = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F191_BMCe_Get"]
SW_serial_BMCe = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F187_BMCe_Get"]
HW_serial_CMC1 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F191_CMC1_Get"]
HW_serial_CMC2 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F191_CMC2_Get"]
HW_serial_CMC3 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F191_CMC3_Get"]
SW_serial_CMC1 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F187_CMC1_Get"]
SW_serial_CMC2 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F187_CMC2_Get"]
SW_serial_CMC3 = Master["[SW_Steuergeraete_Pruefergebnisse]"]["F187_CMC3_Get"]

#Time/Date
timeStamp = Master["[Prueflauf]"]["Startzeit"]
startDate = timeStamp[:timeStamp.find("_")]
startDate = startDate[4:6] + "/" + startDate[6:] + "/" + startDate[:4]
startTime = timeStamp[timeStamp.find("_")+1:].replace("-",":")

#SW Flash Report
preFlashBCMeSW_Check = Master["[SW_Steuergeraete_Ablaufkriterien]"]["PreFlash_BMCSwVers_CheckOK?"]
PostFlashBCMeSW_Check = Master["[SW_Steuergeraete_Ablaufkriterien]"]["PostFlash_BMCSwVers_CheckOK?"]
dataSet_download_Check = Master["[SW_Steuergeraete_Ablaufkriterien]"]["DatensatzdownloadOK?"]
PostDSDL_BCMeSW_Check = Master["[SW_Steuergeraete_Ablaufkriterien]"]["PostDSDL_DSDLSwVers_CheckOK?"]
preFlashCMCSW_Check = Master["[SW_Steuergeraete_Ablaufkriterien]"]["PreFlash_CMCSwVers_CheckOK?"]
PostFlashCMCSW_Check = Master["[SW_Steuergeraete_Ablaufkriterien]"]["PostFlash_CMCSwVers_CheckOK?"]
SW_Check = Master["[SW_Steuergeraete_Ergebnis]"]["0"]

#Battery Status Setpoints Specs
CellTempMin = "{:.2f}".format(float(Master["[Batteriestatus_Sollwertvorgaben]"]["T_ZellMin_Set_Batteriestatus"]))
CellTempMax = "{:.2f}".format(float(Master["[Batteriestatus_Sollwertvorgaben]"]["T_ZellMax_Set_Batteriestatus"]))
CellTempRange = "{:.2f}".format(float(Master["[Batteriestatus_Sollwertvorgaben]"]["dT_ZellMax_Set_Batteriestatus"]))
CellVoltageMin = "{:.2f}".format(float(Master["[Batteriestatus_Sollwertvorgaben]"]["U_ZellMin_Set_Batteriestatus"]))
CellVoltageMax = "{:.2f}".format(float(Master["[Batteriestatus_Sollwertvorgaben]"]["U_ZellMax_Set_Batteriestatus"]))
CellVoltageRange = "{:.2f}".format(float(Master["[Batteriestatus_Sollwertvorgaben]"]["dU_ZellMax_Set_Batteriestatus"]))

#Temp Check
CellTempCheck = Master["[Batteriestatus_Pruefergebnisse]"]["Temperaturwerte iO/niO?"]
CellTemp1 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [0]"]))
CellTemp2 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [1]"]))
CellTemp3 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [2]"]))
CellTemp4 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [3]"]))
CellTemp5 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [4]"]))
CellTemp6 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [5]"]))
CellTemp7 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [6]"]))
CellTemp8 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [7]"]))
CellTemp9 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [8]"]))
CellTemp10 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [9]"]))
CellTemp11 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [10]"]))
CellTemp12 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [11]"]))
CellTemp13 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [12]"]))
CellTemp14 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [13]"]))
CellTemp15 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [14]"]))
CellTemp16 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [15]"]))
CellTemp17 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [16]"]))
CellTemp18 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [17]"]))
CellTemp19 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [18]"]))
CellTemp20 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [19]"]))
CellTemp21 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [20]"]))
CellTemp22 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [21]"]))
CellTemp23 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [22]"]))
CellTemp24 = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["T_Zell_Get[n] [23]"]))
CellTempRange = "{:.2f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["dT_Zell_Calc"]))

#Voltage Check
CellVoltCheck = Master["[Batteriestatus_Pruefergebnisse]"]["Spannungswerte iO/niO?"]
CellVolt1 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [0]"]))
CellVolt2 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [1]"]))
CellVolt3 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [2]"]))
CellVolt4 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [3]"]))
CellVolt5 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [4]"]))
CellVolt6 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [5]"]))
CellVolt7 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [6]"]))
CellVolt8 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [7]"]))
CellVolt9 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [8]"]))
CellVolt10 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [9]"]))
CellVolt11 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [10]"]))
CellVolt12 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [11]"]))
CellVolt13 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [12]"]))
CellVolt14 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [13]"]))
CellVolt15 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [14]"]))
CellVolt16 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [15]"]))
CellVolt17 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [16]"]))
CellVolt18 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [17]"]))
CellVolt19 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [18]"]))
CellVolt20 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [19]"]))
CellVolt21 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [20]"]))
CellVolt22 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [21]"]))
CellVolt23 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [22]"]))
CellVolt24 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [23]"]))
CellVolt25 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [24]"]))
CellVolt26 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [25]"]))
CellVolt27 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [26]"]))
CellVolt28 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [27]"]))
CellVolt29 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [28]"]))
CellVolt30 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [29]"]))
CellVolt31 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [30]"]))
CellVolt32 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [31]"]))
CellVolt33 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [32]"]))
CellVolt34 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [33]"]))
CellVolt35 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [34]"]))
CellVolt36 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [35]"]))
CellVolt37 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [36]"]))
CellVolt38 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [37]"]))
CellVolt39 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [38]"]))
CellVolt40 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [39]"]))
CellVolt41 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [40]"]))
CellVolt42 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [41]"]))
CellVolt43 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [42]"]))
CellVolt44 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [43]"]))
CellVolt45 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [44]"]))
CellVolt46 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [45]"]))
CellVolt47 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [46]"]))
CellVolt48 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [47]"]))
CellVolt49 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [48]"]))
CellVolt50 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [49]"]))
CellVolt51 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [50]"]))
CellVolt52 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [51]"]))
CellVolt53 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [52]"]))
CellVolt54 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [53]"]))
CellVolt55 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [54]"]))
CellVolt56 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [55]"]))
CellVolt57 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [56]"]))
CellVolt58 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [57]"]))
CellVolt59 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [58]"]))
CellVolt60 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [59]"]))
CellVolt61 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [60]"]))
CellVolt62 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [61]"]))
CellVolt63 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [62]"]))
CellVolt64 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [63]"]))
CellVolt65 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [64]"]))
CellVolt66 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [65]"]))
CellVolt67 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [66]"]))
CellVolt68 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [67]"]))
CellVolt69 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [68]"]))
CellVolt70 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [69]"]))
CellVolt71 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [70]"]))
CellVolt72 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [71]"]))
CellVolt73 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [72]"]))
CellVolt74 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [73]"]))
CellVolt75 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [74]"]))
CellVolt76 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [75]"]))
CellVolt77 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [76]"]))
CellVolt78 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [77]"]))
CellVolt79 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [78]"]))
CellVolt80 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [79]"]))
CellVolt81 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [80]"]))
CellVolt82 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [81]"]))
CellVolt83 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [82]"]))
CellVolt84 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [83]"]))
CellVolt85 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [84]"]))
CellVolt86 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [85]"]))
CellVolt87 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [86]"]))
CellVolt88 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [87]"]))
CellVolt89 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [88]"]))
CellVolt90 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [89]"]))
CellVolt91 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [90]"]))
CellVolt92 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [91]"]))
CellVolt93 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [92]"]))
CellVolt94 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [93]"]))
CellVolt95 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [94]"]))
CellVolt96 = "{:.3f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["U_Zell_Get[n] [95]"]))
CellVoltRange = "{:.1f}".format(float(Master["[Batteriestatus_Pruefergebnisse]"]["dU_Zell_Calc"]))
CV_DTC1 = Master["[Batteriestatus_Pruefergebnisse]"]["DTCs_Get [0]"]
CV_DTC2 = Master["[Batteriestatus_Pruefergebnisse]"]["DTCs_Get [1]"]
CV_DTC3 = Master["[Batteriestatus_Pruefergebnisse]"]["DTCs_Get [2]"]
CV_DTC4 = Master["[Batteriestatus_Pruefergebnisse]"]["DTCs_Get [3]"]

#Cell Voltage and Temp Results
CellVoltageMinActual = "{:.3f}".format(float(Master["[Batteriestatus_Ablaufkriterien]"]["U_ZellMin_Get"])/1000)
CellVoltageMaxActual = "{:.3f}".format(float(Master["[Batteriestatus_Ablaufkriterien]"]["U_ZellMax_Get"])/1000)
CellTempMinActual = "{:.3f}".format(float(Master["[Batteriestatus_Ablaufkriterien]"]["T_ZellMin_Get"]))
CellTempMaxActual = "{:.3f}".format(float(Master["[Batteriestatus_Ablaufkriterien]"]["T_ZellMax_Get"]))
TempVoltageResult = Master["[Batteriestatus_Ergebnis]"]["0"]

#Pyrofuse Setpoints
PyroCheckCount = Master["[Pyrofuse_Allgemeine Informationen]"]["uPruefungszaehlerIst"]
PyroMinSet = "{:.5f}".format(float(Master["[Pyrofuse_Sollwertvorgaben]"]["R_PyroMin_Set"]))
PyroMaxSet = "{:.5f}".format(float(Master["[Pyrofuse_Sollwertvorgaben]"]["R_PyroMax_Set"]))

#Pyrofuse Check
PyroCheckResult = "{:.5f}".format(float(Master["[Pyrofuse_Pruefergebnisse]"]["R_Pyro_Mess"]))
PyroCheckStatus = Master["[Pyrofuse_Ergebnis]"]["0"]


#List of values for output file
myExcelData = [battery, partNum, batType, startDate, startTime, i_DTC_Count, initBCMe, initCMC1, initCMC2, initCMC3, initBL_BCMe, initBL_CMC1, initBL_CMC2, initBL_CMC3, batConfigPset, batVehiclePset, batTargetMarketPset, batConfigNamePset, batThermoPset, batNameVehiclePset, batNameTargetMarket, HW_NumCMC1, HW_NumCMC2, HW_NumCMC3,serial_CMC1,serial_CMC2,serial_CMC3,HW_NumBMCe,HW_serial_BMCe,SW_serial_BMCe,HW_serial_CMC1,HW_serial_CMC2,HW_serial_CMC3,SW_serial_CMC1,SW_serial_CMC2,SW_serial_CMC3,\
               preFlashBCMeSW_Check,PostFlashBCMeSW_Check,dataSet_download_Check,PostDSDL_BCMeSW_Check,preFlashCMCSW_Check,PostFlashCMCSW_Check,SW_Check,CellTempMin,CellTempMax,CellTempRange,CellVoltageMin,CellVoltageMax,CellVoltageRange,CellTempCheck,CellTemp1,CellTemp2,CellTemp3,CellTemp4,CellTemp5,CellTemp6,CellTemp7,CellTemp8,CellTemp9,CellTemp10,CellTemp11,CellTemp12,CellTemp13,CellTemp14,CellTemp15,CellTemp16,CellTemp17,CellTemp18,CellTemp19,CellTemp20,CellTemp21,CellTemp22,CellTemp23,CellTemp24,\
               CellTempRange,CellVolt1,CellVolt2,CellVolt3,CellVolt4,CellVolt5,CellVolt6,CellVolt7,CellVolt8,CellVolt9,CellVolt10,CellVolt11,CellVolt12,CellVolt13,CellVolt14,CellVolt15,CellVolt16,CellVolt17,CellVolt18,CellVolt19,CellVolt20,CellVolt21,CellVolt22,CellVolt23,CellVolt24,CellVolt25,CellVolt26,CellVolt27,CellVolt28,CellVolt29,CellVolt30,CellVolt31,CellVolt32,CellVolt33,CellVolt34,CellVolt35,CellVolt36,CellVolt37,CellVolt38,CellVolt39,CellVolt40,CellVolt41,CellVolt42,CellVolt43,CellVolt44,CellVolt45,CellVolt46,CellVolt47,CellVolt48,CellVolt49,CellVolt50,CellVolt51,CellVolt52,CellVolt53,CellVolt54,CellVolt55,CellVolt56,CellVolt57,CellVolt58,CellVolt59,CellVolt60,CellVolt61,CellVolt62,CellVolt63,CellVolt64,CellVolt65,CellVolt66,CellVolt67,CellVolt68,CellVolt69,CellVolt70,CellVolt71,CellVolt72,CellVolt73,CellVolt74,CellVolt75,CellVolt76,CellVolt77,CellVolt78,CellVolt79,CellVolt80,CellVolt81,CellVolt82,CellVolt83,CellVolt84,CellVolt85,CellVolt86,CellVolt87,CellVolt88,CellVolt89,CellVolt90,CellVolt91,CellVolt92,CellVolt93,CellVolt94,CellVolt95,CellVolt96,\
               CellVoltRange,CV_DTC1,CV_DTC2,CV_DTC3,CV_DTC4,CellVoltageMinActual,CellVoltageMaxActual,CellTempMinActual,CellVoltageMaxActual,TempVoltageResult,PyroCheckCount,PyroMinSet,PyroMaxSet,PyroCheckResult,PyroCheckStatus]
myHeaders = ["Battery","Part Number", "Type","Date","Start Time","DTC Count","BCMe Begin SW","CMC1 Begin SW","CMC2 Begin SW","CMC3 Begin SW","BCMe Begin BL","CMC1 Begin BL","CMC2 Begin BL","CMC3 Begin BL","BatConfigPSet","VehiclePSet","TargetMarket","ConfigNamePSet","ThermoPSet","NameVehiclePSet","NameTargetMarket","CMC1 HW","CMC2 HW","CMC3 HW","CMC1 Serial","CMC2 Serial","CMC3 Serial","BCMe HW", "BCMe HW Serial", "BCMe SW Serial", "CMC1 HW Serial","CMC2 HW Serial","CMC3 HW Serial","CMC1 SW Serial","CMC2 SW Serial","CMC3 SW Serial",\
             "preFlashBCMeSW_Check","PostFlashBCMeSW_Check","dataSet_download_Check","PostDSDL_BCMeSW_Check","preFlashCMCSW_Check","PostFlashCMCSW_Check","SW_Check","CellTempMin","CellTempMax","CellTempRange","CellVoltageMin","CellVoltageMax","CellVoltageRange","CellTempCheck","CellTemp1","CellTemp2","CellTemp3","CellTemp4","CellTemp5","CellTemp6","CellTemp7","CellTemp8","CellTemp9","CellTemp10","CellTemp11","CellTemp12","CellTemp13","CellTemp14","CellTemp15","CellTemp16","CellTemp17","CellTemp18","CellTemp19","CellTemp20","CellTemp21","CellTemp22","CellTemp23","CellTemp24",\
             "CellTempRange","CellVolt1","CellVolt2","CellVolt3","CellVolt4","CellVolt5","CellVolt6","CellVolt7","CellVolt8","CellVolt9","CellVolt10","CellVolt11","CellVolt12","CellVolt13","CellVolt14","CellVolt15","CellVolt16","CellVolt17","CellVolt18","CellVolt19","CellVolt20","CellVolt21","CellVolt22","CellVolt23","CellVolt24","CellVolt25","CellVolt26","CellVolt27","CellVolt28","CellVolt29","CellVolt30","CellVolt31","CellVolt32","CellVolt33","CellVolt34","CellVolt35","CellVolt36","CellVolt37","CellVolt38","CellVolt39","CellVolt40","CellVolt41","CellVolt42","CellVolt43","CellVolt44","CellVolt45","CellVolt46","CellVolt47","CellVolt48","CellVolt49","CellVolt50","CellVolt51","CellVolt52","CellVolt53","CellVolt54","CellVolt55","CellVolt56","CellVolt57","CellVolt58","CellVolt59","CellVolt60","CellVolt61","CellVolt62","CellVolt63","CellVolt64","CellVolt65","CellVolt66","CellVolt67","CellVolt68","CellVolt69","CellVolt70","CellVolt71","CellVolt72","CellVolt73","CellVolt74","CellVolt75","CellVolt76","CellVolt77","CellVolt78","CellVolt79","CellVolt80","CellVolt81","CellVolt82","CellVolt83","CellVolt84","CellVolt85","CellVolt86","CellVolt87","CellVolt88","CellVolt89","CellVolt90","CellVolt91","CellVolt92","CellVolt93","CellVolt94","CellVolt95","CellVolt96",\
             "CellVoltRange","CV_DTC1","CV_DTC2","CV_DTC3","CV_DTC4","CellVoltageMinActual","CellVoltageMaxActual","CellTempMinActual","CellVoltageMaxActual","TempVoltageResult","PyroCheckCount","PyroMinSet","PyroMaxSet","PyroCheckResult","PyroCheckStatus"]
count = 0

for header in myHeaders:
    count = count + 1
    sheet.cell(1, count).value = header

count = 0
for myVals in myExcelData:
    count = count + 1
    sheet.cell(2,count).value =myVals

#Format for time
sheet.cell(2,5).number_format = "h:mm:ss AM/PM"

for i in range(len(myExcelData)):
    print(myHeaders[i] + " : " +myExcelData[i])

wb.save(filename="hello_world.xlsx")