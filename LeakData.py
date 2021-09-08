from openpyxl import Workbook
import os

#Function for getting data from file
def getLeak(myFile):
    input_file = open(myFile, 'r')

    # Variable Setup
    # Master dictionary
    Master = {}

    # Loop through each line in file
    while input_file:
        # Read next line
        currentLine = input_file.readline()
        # Check if EOF
        if currentLine == "":
            break
        # Identify header (will be separate dictionary entry in master dictionary
        if currentLine[0] == "[":
            header = currentLine.strip()
            # Create a blank dictionary to temporarily save to before adding to master
            myDict = {}
            dictName = currentLine

            # Index to next line
            currentLine = input_file.readline()

            while currentLine.strip() != "":

                # If there is a ";" in the string
                if ";" in currentLine.strip():
                    data = currentLine[currentLine.find('"') + 1:].strip()
                    data = data[:-1]
                    # Create list of data/value pair
                    myData = data.split(";")
                    myDict[myData[0]] = myData[1]
                # If there is NOT a ";" in the string
                else:
                    #      0 = "IO"
                    data = currentLine
                    myData = currentLine.split("=")
                    myDict[myData[0].strip(" ")] = myData[1].strip().strip(" ")

                # Index to the next line at end of loop
                currentLine = input_file.readline()

            Master[header] = myDict

    input_file.close()
    ###############################################
    #Check if valid leak file
    if not Master["[Pruefstand]"]["Pruefstandstyp"] == "Dichtigkeit":
        return False
    if "[Pruefling]" not in Master.keys():
        print(input_file.name.split('\\')[-1])

    # General Data
    try:
        battery, partNum, batType = "","",""
        battery = Master["[Pruefling]"]["Seriennummer"].replace("288 TX6","").upper().replace("X","").replace(" ","")
        partNum = Master["[Pruefling]"]["Teilenummer"]
        batType = Master["[Pruefling]"]["Batterietyp"]
    except:
        pass

    # Time/Date
    timeStamp = Master["[Prueflauf]"]["Startzeit"]
    startDate = timeStamp[:timeStamp.find("_")]
    startDate = startDate[4:6] + "/" + startDate[6:] + "/" + startDate[:4]
    startTime = timeStamp[timeStamp.find("_") + 1:].replace("-", ":")

    # Leak Information
    if "[pPosZSB_Pruefergebnisse]" in Master.keys():
        overPressure = round(float(Master["[pPosZSB_Pruefergebnisse]"]["Q_ZSB2_Mess"]), 2)
        resultOver = Master["[pPosZSB_Ergebnis]"]["0"].strip('"')
    else:
        overPressure = ""
        resultOver = ""

    if "[pNegZSB_pPosKS_Pruefergebnisse]" in Master.keys():
        underPressure = round(float(Master["[pNegZSB_pPosKS_Pruefergebnisse]"]["Q_ZSB1_Mess"]), 2)
        coolPressure = round(float(Master["[pNegZSB_pPosKS_Pruefergebnisse]"]["Q_Kuehl_Mess"]), 2)
        resultUnder = Master["[pNegZSB_pPosKS_Ergebnis]"]["0"].strip('"')
    else:
        coolPressure = ""
        underPressure = ""
        resultUnder = ""

    resultFinal = Master["[Ergebnis]"]["0"].strip('"')


        # List of values for output file
    myExcelData = [startDate, startTime, batType, battery, partNum, coolPressure, underPressure, resultUnder, overPressure,resultOver,
                   resultFinal]

    return myExcelData

#Directory for leak test data
myLeakDir = r'\\na.vwg\chattanooga\Dept\CA\50_ELECTRIC\10 Battery Specialist\210_Leak\Results\Results'

#Setup Excel workbook
wb = Workbook()
sheet = wb.active

#Setup headers
myHeaders = ["Date", "Timestamp", "Model", "Serial", "Part Number","Cooling Leak Rate\n(cc/m)","UnderPressure Leak Rate\n(cc/m)","UnderResult", "OverPressure Leak Rate\n(cc/m)", "OverResult",
             "Overall Result", "Filename"]
for i in range(len(myHeaders)):
    sheet.cell(1,i+1).value = myHeaders[i]

#Populate excel sheet w/ test data
lineNum = 2
for testData in os.listdir(myLeakDir):
    #Get next filename
    myLeak = myLeakDir + "\\" + testData
    #Call file parser function
    myExcelData = getLeak(myLeak)
    if myExcelData == False:
        continue
    count = 0
    for myVals in myExcelData:
        count = count + 1
        sheet.cell(lineNum, count).value = myVals
    sheet.cell(lineNum,count+1).value = testData
    # Format for time
    sheet.cell(lineNum, 5).number_format = "h:mm:ss AM/PM"
    lineNum += 1
    #print(myExcelData[3])
    wb.save(filename="leak_data.xlsx")

#Adjust cell columns
for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[column_cells[0].column_letter].width = length

wb.save(filename="leak_data.xlsx")







#myLeak = r'\\vwoachsfile01\assembly\Departments\Assembly Launch\Battery Plant Status\Pilot Hall\IOL_EOL_Leak Docs\PVS\288 TX6PVS0324_2021_07_12_213822.ini'
